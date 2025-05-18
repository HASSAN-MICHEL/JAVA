import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import io
import os
from datetime import datetime


stage = pd.read_excel("STAGE.xlsx")


# Configuration initiale
sns.set_style('whitegrid')
plt.rcParams['figure.dpi'] = 300
plt.rcParams['savefig.dpi'] = 300

# Création du classeur Excel
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Résumé"

# Suppression des feuilles inutiles
for sheet in wb.sheetnames[1:]:
    del wb[sheet]

# Fonction pour sauvegarder les graphiques
def save_plot_to_excel(worksheet, fig, cell):
    """Sauvegarde un graphique matplotlib dans une feuille Excel"""
    img_buffer = io.BytesIO()
    fig.savefig(img_buffer, format='png', bbox_inches='tight')
    img_buffer.seek(0)
    img = Image(img_buffer)
    worksheet.add_image(img, cell)
    plt.close(fig)

# Fonction pour ajouter un DataFrame avec style
def add_dataframe(worksheet, df, title, start_cell):
    """Ajoute un DataFrame à une feuille Excel avec un titre"""
    worksheet[start_cell] = title
    for r in dataframe_to_rows(df, index=True, header=True):
        worksheet.append(r)

# Chargement des données
stage = pd.read_excel("STAGE.xlsx")
stage['Année_vente'] = stage['Date_vente'].dt.year

# 1. FEUILLE DE RÉSUMÉ

ws_summary['A1'] = "Rapport d'Analyse des Ventes"
ws_summary['A2'] = f"Généré le {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

metrics = [
    ("Chiffre d'affaires total", f"{stage['Prix_vente (€)'].sum():,.2f}€"),
    ("Prix moyen de vente", f"{stage['Prix_vente (€)'].mean():,.2f}€"),
    ("Remise moyenne", f"{stage['Remise (€)'].mean():,.2f}€"),
    ("Score moyen de satisfaction", f"{stage['Score_satisfaction_client (1-10)'].mean():.2f}/10"),
    ("Délai moyen de livraison", f"{stage['Délai_livraison (jours)'].mean():.1f} jours")
]

for i, (label, value) in enumerate(metrics, start=3):
    ws_summary[f'A{i}'] = label
    ws_summary[f'B{i}'] = value


# 2. STATISTIQUES DESCRIPTIVES
# =====================================================================
ws_stats = wb.create_sheet("Statistiques")

# Informations générales
add_dataframe(ws_stats, stage.describe(include='all').transpose(), "Statistiques descriptives", "A1")

# Valeurs manquantes
missing_values = stage.isnull().sum().to_frame("Valeurs manquantes")
add_dataframe(ws_stats, missing_values, "Valeurs manquantes", "A20")

# 3. ANALYSE PAR PAYS

ws_pays = wb.create_sheet("Ventes par Pays")

vente_pays = stage.groupby('Pays_vente')['Prix_vente (€)'].sum().reset_index()
vente_pays = vente_pays.sort_values('Prix_vente (€)', ascending=False)
add_dataframe(ws_pays, vente_pays, "Ventes par pays (CA total)", "A1")

# Graphique ventes par pays
fig, ax = plt.subplots(figsize=(12, 6))
sns.barplot(data=vente_pays, x='Pays_vente', y='Prix_vente (€)', ax=ax, palette='viridis')
ax.set_title('Chiffre d\'affaires par pays')
ax.set_ylabel('CA (€)')
ax.set_xlabel('Pays')
plt.xticks(rotation=45)
save_plot_to_excel(ws_pays, fig, "D1")

# 4. ANALYSE PAR MODÈLE

ws_modele = wb.create_sheet("Ventes par Modèle")

ventes_par_modele = stage.groupby(['Marque', 'Modèle']).agg({
    'ID_vente': 'count',
    'Prix_vente (€)': ['mean', 'sum']
}).sort_values(('ID_vente', 'count'), ascending=False)
ventes_par_modele.columns = ['Nombre_ventes', 'Prix_moyen', 'CA_total']
add_dataframe(ws_modele, ventes_par_modele.head(10), "Top 10 modèles les plus vendus", "A1")

# Graphique top 10 modèles
fig, ax = plt.subplots(figsize=(12, 6))
ventes_par_modele.head(10)['Nombre_ventes'].plot(kind='bar', ax=ax)
ax.set_title('Top 10 des modèles les plus vendus')
ax.set_ylabel('Nombre de ventes')
plt.xticks(rotation=45)
save_plot_to_excel(ws_modele, fig, "D20")


# 5. ANALYSE PAR MARQUE

ws_marque = wb.create_sheet("Ventes par Marque")

vente_marques = stage['Marque'].value_counts().reset_index()
vente_marques.columns = ['Marque', 'Nombre_ventes']
add_dataframe(ws_marque, vente_marques, "Ventes par marque", "A1")

# Graphique ventes par marque
fig, ax = plt.subplots(figsize=(12, 8))
ventes_par_marque = stage['Marque'].value_counts().sort_values()
bars = ax.barh(ventes_par_marque.index, ventes_par_marque.values, color='skyblue')

for bar in bars:
    width = bar.get_width()
    ax.text(width + 0.5, bar.get_y() + bar.get_height()/2,
            f'{int(width)} ({width/len(stage)*100:.1f}%)',
            va='center', ha='left')

ax.set_title('Nombre de ventes par marque')
ax.set_xlabel('Nombre de ventes')
ax.set_ylabel('Marque')
ax.grid(axis='x', linestyle='--', alpha=0.7)
save_plot_to_excel(ws_marque, fig, "D1")

# Graphique prix par marque et carburant
fig, ax = plt.subplots(figsize=(14, 8))
top_marques = stage['Marque'].value_counts().head(10).index
sns.boxplot(x='Marque', y='Prix_vente (€)', hue='Type_carburant', 
            data=stage[stage['Marque'].isin(top_marques)], ax=ax)
ax.set_title('Prix moyen par marque et type de carburant')
plt.xticks(rotation=45)
save_plot_to_excel(ws_marque, fig, "D30")


# 6. ÉVOLUTION ANNUELLE

ws_evolution = wb.create_sheet("Evolution Annuelle")

evolution_annuelle = stage.groupby('Année_vente').agg({
    'ID_vente': 'count',
    'Prix_vente (€)': 'sum'
}).rename(columns={'ID_vente': 'Nombre_ventes', 'Prix_vente (€)': 'CA'})
add_dataframe(ws_evolution, evolution_annuelle, "Evolution annuelle", "A1")

# Graphique évolution annuelle
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 12))

# Nombre de ventes
bars1 = ax1.bar(evolution_annuelle.index.astype(str), evolution_annuelle['Nombre_ventes'], color='skyblue')
ax1.set_title('Nombre de ventes par année')
ax1.set_ylabel('Nombre de ventes')
ax1.grid(axis='y', linestyle='--', alpha=0.7)

# Chiffre d'affaires
bars2 = ax2.bar(evolution_annuelle.index.astype(str), evolution_annuelle['CA'], color='salmon')
ax2.set_title('Chiffre d\'affaires par année')
ax2.set_ylabel('CA (€)')
ax2.grid(axis='y', linestyle='--', alpha=0.7)

plt.tight_layout()
save_plot_to_excel(ws_evolution, fig, "D1")


# 7. ANALYSE CROISÉE

ws_croisee = wb.create_sheet("Analyse Croisée")

# Analyse croisée complète
analyse_croisee = stage.groupby(['Marque', 'Année_modèle', 'Type_carburant']).agg({
    'ID_vente': 'count',
    'Prix_vente (€)': 'mean',
    'Client_type': lambda x: x.value_counts().index[0],
    'Score_satisfaction_client (1-10)': 'mean'
}).rename(columns={
    'ID_vente': 'Nombre_ventes',
    'Prix_vente (€)': 'Prix_moyen',
    'Score_satisfaction_client (1-10)': 'Satisfaction_moyenne'
}).reset_index()

add_dataframe(ws_croisee, analyse_croisee.head(10), "Top 10 combinaisons marque/modèle/carburant", "A1")

# Graphique top combinaisons
top10_combinaisons = analyse_croisee.sort_values('Nombre_ventes', ascending=False).head(10)
fig, ax = plt.subplots(figsize=(14, 8))
sns.barplot(data=top10_combinaisons, x='Nombre_ventes', y='Marque', hue='Type_carburant', palette='viridis', ax=ax)
ax.set_title('Top 10 des combinaisons Marque/Carburant les plus vendues')
ax.set_xlabel('Nombre de ventes')
ax.set_ylabel('Marque')
plt.legend(title='Type carburant', bbox_to_anchor=(1.05, 1), loc='upper left')
save_plot_to_excel(ws_croisee, fig, "D20")

# Heatmap
pivot_table = pd.pivot_table(analyse_croisee,
                            values='Nombre_ventes',
                            index='Marque',
                            columns='Type_carburant',
                            aggfunc='sum',
                            fill_value=0)

fig, ax = plt.subplots(figsize=(16, 10))
sns.heatmap(pivot_table, annot=True, fmt='g', cmap='YlGnBu', linewidths=.5, ax=ax)
ax.set_title('Répartition des ventes par marque et type de carburant')
ax.set_xlabel('Type de carburant')
ax.set_ylabel('Marque')
save_plot_to_excel(ws_croisee, fig, "D50")


# 8. ANALYSE CLIENTS

ws_clients = wb.create_sheet("Analyse Clients")

# Répartition clients par marque
client_marque = stage.groupby(['Marque', 'Client_type']).size().unstack()
add_dataframe(ws_clients, client_marque, "Répartition des clients par marque", "A1")

# Graphique clients par marque
fig, ax = plt.subplots(figsize=(14, 8))
client_marque.plot(kind='bar', stacked=True, colormap='Paired', ax=ax)
ax.set_title('Répartition des types de clients par marque')
ax.set_xlabel('Marque')
ax.set_ylabel('Nombre de clients')
plt.xticks(rotation=45)
plt.legend(title='Type de client', bbox_to_anchor=(1.05, 1), loc='upper left')
save_plot_to_excel(ws_clients, fig, "D20")


# 9. CORRÉLATIONS

ws_corr = wb.create_sheet("Corrélations")

# Matrice de corrélation
numerical_cols = ['Prix_vente (€)', 'Année_modèle', 'Kilométrage (km)', 'Garantie (mois)', 
                 'Score_satisfaction_client (1-10)', 'Remise (€)', 'Délai_livraison (jours)']
corr_matrix = stage[numerical_cols].corr()
add_dataframe(ws_corr, corr_matrix, "Matrice de corrélation", "A1")

# Graphique corrélation
fig, ax = plt.subplots(figsize=(12, 8))
sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', center=0, ax=ax)
ax.set_title('Matrice de corrélation')
save_plot_to_excel(ws_corr, fig, "D20")
# SAUVEGARDE DU FICHIER

output_dir = "Rapports_Analyse"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_path = os.path.join(output_dir, f"Rapport_Complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
wb.save(output_path)

print(f"Rapport généré avec succès : {output_path}")
